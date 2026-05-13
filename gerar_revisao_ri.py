import requests, pandas as pd, unicodedata, math, time, io, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def norm(t):
    if not isinstance(t, str): return ''
    t = unicodedata.normalize('NFKD', t)
    t = ''.join(c for c in t if not unicodedata.combining(c))
    return t.strip().upper()

def haversine(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1))*math.cos(math.radians(lat2))*math.sin(dlon/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

print("Baixando coordenadas IBGE...")
r = requests.get('https://raw.githubusercontent.com/kelvins/municipios-brasileiros/main/csv/municipios.csv', timeout=15)
coords = pd.read_csv(io.StringIO(r.text))
coords['nome_norm'] = coords['nome'].apply(norm)
uf_map = {
    41:'PR', 13:'AM', 35:'SP', 51:'MT', 11:'RO', 12:'AC', 14:'RR', 15:'PA', 16:'AP', 17:'TO',
    21:'MA', 22:'PI', 23:'CE', 24:'RN', 25:'PB', 26:'PE', 27:'AL', 28:'SE', 29:'BA', 31:'MG',
    32:'ES', 33:'RJ', 42:'SC', 43:'RS', 50:'MS', 52:'GO', 53:'DF'
}
coords['uf_sigla'] = coords['codigo_uf'].map(uf_map)
coords['codigo_ibge'] = coords['codigo_ibge'].astype(str)

BASE = 'https://justicaabertaapi.cnj.jus.br/v1/api'

df_cache = pd.read_csv('data/cache/cartorios_cnj.csv')
df_cache['municipio_norm'] = df_cache['municipio'].apply(norm)
cidades_com_ri = set(df_cache['municipio_norm'] + '|' + df_cache['uf'])
ri_por_uf = df_cache[['municipio_norm', 'municipio', 'uf']].drop_duplicates(subset=['municipio_norm', 'uf'])

estados = ['PR', 'AM', 'SP', 'MT']
sem_ri = []
print("Buscando municípios sem RI...")
for uf in estados:
    r2 = requests.get(f'{BASE}/cidades/listar/{uf}', timeout=15)
    for c in r2.json():
        if norm(c['nome']) + '|' + uf not in cidades_com_ri:
            sem_ri.append({
                'municipio': c['nome'],
                'municipio_norm': norm(c['nome']),
                'uf': uf,
                'codigo_ibge': str(c.get('codigo_ibge', ''))
            })
    print(f"  {uf}: {sum(1 for x in sem_ri if x['uf']==uf)} sem RI")
    time.sleep(0.3)

df_sem_ri = pd.DataFrame(sem_ri)
df_sem_ri = df_sem_ri.merge(coords[['codigo_ibge', 'latitude', 'longitude']], on='codigo_ibge', how='left')

coords_ri = coords[coords['uf_sigla'].isin(estados)].copy()
coords_ri = coords_ri.merge(
    ri_por_uf, left_on=['nome_norm', 'uf_sigla'], right_on=['municipio_norm', 'uf'], how='inner'
)

print("Calculando distâncias...")
resultados = []
for _, row in df_sem_ri.iterrows():
    candidatos = coords_ri[coords_ri['uf_sigla'] == row['uf']].copy()

    if candidatos.empty or pd.isna(row.get('latitude')):
        resultados.append({
            'municipio': row['municipio'], 'uf': row['uf'],
            'ri_sugerido': 'SEM COORDENADA', 'distancia_km': None,
            'municipio_ri_confirmado': '', 'observacao': 'Verificar manualmente'
        })
        continue

    candidatos = candidatos.copy()
    candidatos['dist'] = candidatos.apply(
        lambda c: haversine(row['latitude'], row['longitude'], c['latitude'], c['longitude']), axis=1
    )
    mais_proximo = candidatos.nsmallest(1, 'dist').iloc[0]
    dist = round(mais_proximo['dist'], 1)

    ri_nome = mais_proximo.get('municipio_y') or mais_proximo.get('municipio_norm', '')
    status = 'Confirmar' if dist <= 80 else 'Revisar - distancia alta'

    resultados.append({
        'municipio': row['municipio'], 'uf': row['uf'],
        'ri_sugerido': ri_nome, 'distancia_km': dist,
        'municipio_ri_confirmado': '', 'observacao': status
    })

df_result = pd.DataFrame(resultados).sort_values(['uf', 'municipio']).reset_index(drop=True)

# Remover casos já no override (Tapejara e Nova Maringá)
try:
    df_override = pd.read_csv('data/overrides/municipio_ri_override.csv')
    ja_mapeados = set(df_override['municipio_norm'].apply(norm) + '|' + df_override['uf_sigla'])
    df_result = df_result[~(df_result['municipio'].apply(norm) + '|' + df_result['uf']).isin(ja_mapeados)]
except Exception:
    pass

print(f"Total para revisão: {len(df_result)}")

# Excel formatado
os.makedirs('data/overrides', exist_ok=True)
out_path = 'data/overrides/municipios_sem_ri_para_revisao.xlsx'

wb = Workbook()
ws = wb.active
ws.title = "Para Revisar"

header_font   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
header_fill   = PatternFill('solid', start_color='1F4E79')
yellow_fill   = PatternFill('solid', start_color='FFF2CC')
green_fill    = PatternFill('solid', start_color='E2EFDA')
orange_fill   = PatternFill('solid', start_color='FCE4D6')
thin          = Side(style='thin', color='BFBFBF')
bord          = Border(left=thin, right=thin, top=thin, bottom=thin)
center_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align    = Alignment(horizontal='left', vertical='center')

# Linha de instrucoes
ws.merge_cells('A1:F1')
ws['A1'] = (
    'INSTRUCOES: Revise "RI Sugerido". Se correto, deixe "RI Confirmado" em branco. '
    'Se errado, preencha com o municipio correto. '
    'Amarelo = OK para confirmar. Laranja = distancia alta, revise com atencao.'
)
ws['A1'].font = Font(name='Arial', bold=True, size=9, color='1F4E79')
ws['A1'].fill = PatternFill('solid', start_color='DEEAF1')
ws['A1'].alignment = Alignment(wrap_text=True, vertical='center')
ws.row_dimensions[1].height = 35

headers    = ['Municipio', 'UF', 'RI Sugerido (distancia)', 'Distancia (km)', 'RI Confirmado\n(preencher se errado)', 'Status']
col_widths = [30, 6, 32, 15, 32, 24]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = bord
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[2].height = 30

for i, row in enumerate(df_result.itertuples(), 3):
    dist_val = row.distancia_km
    is_alta  = dist_val is not None and dist_val > 80
    row_fill = orange_fill if is_alta else yellow_fill

    values = [row.municipio, row.uf, row.ri_sugerido,
              round(dist_val, 1) if dist_val else '', '', row.observacao]

    for col, val in enumerate(values, 1):
        c = ws.cell(i, col, val)
        c.font   = Font(name='Arial', size=9)
        c.border = bord
        c.fill   = green_fill if col == 5 else row_fill
        c.alignment = left_align if col in (1, 3, 5) else center_align

    ws.row_dimensions[i].height = 15

# Aba Resumo
ws2 = wb.create_sheet('Resumo')
resumo_headers = ['Estado', 'Total sem RI', 'Sugestao <= 80km', 'Sugestao > 80km']
for col, h in enumerate(resumo_headers, 1):
    c = ws2.cell(1, col, h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = bord
    ws2.column_dimensions[get_column_letter(col)].width = 20

for i, uf in enumerate(['PR', 'AM', 'SP', 'MT'], 2):
    sub = df_result[df_result['uf'] == uf]
    ok  = len(sub[(sub['distancia_km'].fillna(999) <= 80)])
    rev = len(sub[(sub['distancia_km'].fillna(0)   > 80)])
    for col, val in enumerate([uf, len(sub), ok, rev], 1):
        c = ws2.cell(i, col, val)
        c.font      = Font(name='Arial', size=9)
        c.border    = bord
        c.alignment = center_align

wb.save(out_path)
print(f"\nArquivo salvo: {out_path}")
print("\nDistribuicao por estado:")
print(df_result['uf'].value_counts().to_string())
print("\nStatus das sugestoes:")
print(df_result['observacao'].value_counts().to_string())
